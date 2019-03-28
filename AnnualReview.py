from selenium import webdriver
from Login import MainLogin
import openpyxl
import json

class AnnualReview(MainLogin):
    def __init__(self):
        # self.driver = webdriver.Chrome(executable_path=r"C:\chromedriver\chromedriver.exe")
        # self.driver.maximize_window()
        # self.login("https://devlife.gaditek-hrms.sg.plesk-server.com/annual-review/questions")
        self.create_question()

    def create_question(self):

        filename = "C:\\Users\mojis.mehdi\PycharmProjects\Life\Ques.JSON"
        with open(filename) as f:
            pop_data = json.load(f)
            print(pop_data)
        questions = []
        # wb = openpyxl.load_workbook('C:\\Users\Gaditek\PycharmProjects\Life\Question.xlsx')
        # sheet = wb.active
        # # get max row count
        # max_row = sheet.max_row
        # # get max column count
        # max_column = sheet.max_column
        # # iterate over all cells
        # # iterate over all rows
        # for i in range(2, max_row + 1):
        #     # iterate over all columns
        #     for j in range(1, max_column + 1):
        #         # get particular cell value
        #         cell_obj = sheet.cell(row=i, column=j)
        #         # print cell value
        #         cell_obj_heading = sheet.cell(1, j).value
        #         questions.append({
        #             cell_obj_heading: cell_obj.value
        #         });
        # print(questions)

        # for i in range(len(questions)):
        #     print(questions[i])
        #     if questions[i].Question
        #     i += 1
        #     print('\n')


annualreview = AnnualReview()
